#!/usr/bin/env python3
"""Build a self-contained variables-browser HTML and Excel workbook.

Pipeline:
    1. Scan ``./data`` for ``<dataset>_register_codebook.yaml`` and
       ``<dataset>_register_meta.yaml`` files.
    2. For each dataset, load and normalise variable codebooks + metadata.
       Dataset descriptions prefer a companion Markdown file
       (``<dataset>_register_meta.md``); if absent, fall back to the YAML
       ``info:`` field.
    3. Apply ignore / grouping rules from
       ``./config/variables_config.yaml``.
    4. Render ``./templates/index.html.j2`` to
       ``./dist/variables_browser.html`` with embedded per-file
       provenance (SHA-256) and a combined digest.
    5. Export formatted ``./dist/variables_browser.xlsx`` for offline use.
"""

import argparse
import base64
import fnmatch
import hashlib
import json
import logging
import os
import re
import shutil
import subprocess
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import markdown as markdown_lib
import yaml
from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl import Workbook
from openpyxl.packaging.core import DocumentProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

# Type alias for YAML-derived dictionaries whose values are
# heterogeneous and not worth spelling out at every call site.
_YamlDict = dict[str, Any]

# ──────────────────────────────────────────────────────────────────────
# Paths (defaults — ``--output-dir`` overrides DIST_DIR at runtime)
# ──────────────────────────────────────────────────────────────────────

ROOT_DIR = Path(__file__).resolve().parent
DATA_DIR = ROOT_DIR / "data"
CONTENT_DIR = ROOT_DIR / "content"
CONFIG_DIR = ROOT_DIR / "config"
TEMPLATES_DIR = ROOT_DIR / "templates"
DIST_DIR = ROOT_DIR / "dist"

DATASETS_CONFIG = CONFIG_DIR / "datasets_config.yaml"
VARIABLES_CONFIG = CONFIG_DIR / "variables_config.yaml"

TEMPLATE_NAME = "index.html.j2"
OUTPUT_HTML_NAME = "variables_browser.html"

INTRO_MD = CONTENT_DIR / "intro.md"
FOOTER_MD = CONTENT_DIR / "footer.md"
LOGO_PATH = CONTENT_DIR / "logo.png"
STATIC_DIR = ROOT_DIR / "static"

# Recognised keys inside each ``groups:`` entry in variables_config.yaml.
# Used for typo-detection warnings at load time.
_KNOWN_GROUP_KEYS = frozenset({
    "pattern",
    "source_variable_name_grouped",
    "label",
    "notes",
    "csv_expand",
    "category_strategy",
    "categories_override",
    "priority",
})


# ──────────────────────────────────────────────────────────────────────
# Hashing utilities
# ──────────────────────────────────────────────────────────────────────

def _sha256_bytes(data: bytes) -> str:
    """Return the hex SHA-256 digest of *data*.

    Args:
        data: Raw bytes to hash.

    Returns:
        Hex-encoded SHA-256 digest string.
    """
    return hashlib.sha256(data).hexdigest()


def _sha256_file(path: Path) -> str:
    """Return the hex SHA-256 digest of a file, read in chunks.

    Reading in 64 KiB chunks avoids loading very large files entirely
    into memory.

    Args:
        path: Filesystem path to the file.

    Returns:
        Hex-encoded SHA-256 digest string.
    """
    digest = hashlib.sha256()
    with path.open("rb") as fh:
        while chunk := fh.read(65_536):
            digest.update(chunk)
    return digest.hexdigest()


# ──────────────────────────────────────────────────────────────────────
# YAML / Markdown helpers
# ──────────────────────────────────────────────────────────────────────

def _read_yaml(path: Path) -> Any:
    """Load a YAML file using the safe loader.

    Args:
        path: Filesystem path to a ``.yaml`` file.

    Returns:
        The parsed YAML content (typically a ``dict`` or ``list``).
    """
    with path.open("r", encoding="utf-8") as fh:
        return yaml.safe_load(fh)


def _md_to_html(
    text: str,
    *,
    preserve_newlines: bool = False,
) -> str:
    """Convert Markdown text to HTML.

    Uses the *extra* and *admonition* extensions.  When
    *preserve_newlines* is ``True`` the *nl2br* extension is added so
    that single newlines inside paragraphs become ``<br>`` tags (useful
    for dataset description cards where authors expect line-break
    fidelity).

    Args:
        text: Raw Markdown string.
        preserve_newlines: If ``True``, add the ``nl2br`` extension.

    Returns:
        Rendered HTML string, or ``""`` on empty input.
    """
    if not text:
        return ""
    extensions = ["extra", "admonition"]
    if preserve_newlines:
        extensions.append("nl2br")
    try:
        return markdown_lib.markdown(text, extensions=extensions)
    except Exception:
        logger.warning(
            "Markdown conversion failed; using plain-text fallback.",
        )
        paragraphs = [
            p.strip()
            for p in re.split(r"\n\s*\n", text.strip())
            if p.strip()
        ]
        return "".join(
            f"<p>{_escape_html(p)}</p>" for p in paragraphs
        )


def _escape_html(text: str) -> str:
    """Escape the five HTML-significant characters.

    Args:
        text: Unescaped string.

    Returns:
        String safe for inclusion in HTML element content.
    """
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&#39;")
    )


# ──────────────────────────────────────────────────────────────────────
# Dataset identification
# ──────────────────────────────────────────────────────────────────────

def _dataset_id_from_filename(path: Path) -> str:
    """Extract a dataset identifier from a data-file name.

    Examples:
        ``ps_cancer_register_codebook.yaml``  -> ``ps_cancer``
        ``ps_cause-of-death_register_meta.yaml`` -> ``ps_cause-of-death``

    Args:
        path: Path to a ``*_register_codebook.yaml`` or
            ``*_register_meta.yaml`` file.

    Returns:
        Dataset identifier string.
    """
    name = path.name
    for suffix in ("_register_codebook.yaml", "_register_meta.yaml"):
        if name.endswith(suffix):
            return name[: -len(suffix)]
    # Fallback: strip known substrings from the stem.
    stem = path.stem
    stem = stem.replace("_register_codebook", "")
    stem = stem.replace("_register_meta", "")
    return stem


# ──────────────────────────────────────────────────────────────────────
# variables_config.yaml handling
# ──────────────────────────────────────────────────────────────────────

def _compile_ignore_patterns(
    patterns: list[str] | None,
) -> list[re.Pattern[str]]:
    """Compile name-ignore patterns as regular expressions.

    Each pattern is first attempted as a raw regex.  If that fails
    (e.g. the author supplied a glob like ``*.tmp``), it is converted
    via :func:`fnmatch.translate` and compiled as a regex instead.

    Args:
        patterns: Raw pattern strings from the config file.

    Returns:
        List of compiled :class:`re.Pattern` objects.
    """
    compiled: list[re.Pattern[str]] = []
    for raw in patterns or []:
        pattern_str = (raw or "").strip()
        if not pattern_str:
            continue
        try:
            compiled.append(re.compile(pattern_str))
        except re.error:
            logger.warning(
                "Regex compilation failed for '%s'; "
                "treating as glob pattern.",
                pattern_str,
            )
            compiled.append(
                re.compile(fnmatch.translate(pattern_str)),
            )
    return compiled


def _load_variables_config() -> tuple[
    dict[str, _YamlDict],
    list[str],
    list[re.Pattern[str]],
    list[str],
    list[str],
]:
    """Load variable filtering and grouping rules.

    Reads ``config/variables_config.yaml`` and returns five values
    consumed by the assembly pipeline.

    Returns:
        A 5-tuple of:
            - groups_config: Mapping of group-ID -> group definition.
            - ignore_names: Exact variable names to drop.
            - ignore_name_patterns: Compiled regex patterns; matching
              variable names are dropped.
            - ignore_tags: Variables carrying any of these tags are
              dropped.
            - ignore_categories: Variables in any of these categories
              are dropped.
    """
    if not VARIABLES_CONFIG.exists():
        return {}, [], [], [], []

    config = _read_yaml(VARIABLES_CONFIG) or {}
    groups_config: dict[str, _YamlDict] = config.get("groups") or {}

    # Warn about unrecognised keys (catches typos like "patern").
    for group_id, group_def in groups_config.items():
        if not isinstance(group_def, dict):
            continue
        unknown = set(group_def.keys()) - _KNOWN_GROUP_KEYS
        if unknown:
            logger.warning(
                "Unrecognised key(s) %s in group '%s' "
                "(variables_config.yaml). Recognised: %s",
                sorted(unknown),
                group_id,
                sorted(_KNOWN_GROUP_KEYS),
            )

    ignore_cfg = config.get("ignore") or {}
    ignore_names: list[str] = ignore_cfg.get("names") or []
    ignore_name_patterns = _compile_ignore_patterns(
        ignore_cfg.get("name_patterns") or [],
    )
    ignore_tags: list[str] = ignore_cfg.get("tags") or []
    ignore_categories: list[str] = (
        ignore_cfg.get("categories") or []
    )

    return (
        groups_config,
        ignore_names,
        ignore_name_patterns,
        ignore_tags,
        ignore_categories,
    )


# ──────────────────────────────────────────────────────────────────────
# Codebook / meta ingestion
# ──────────────────────────────────────────────────────────────────────

def _discover_datasets() -> dict[str, dict[str, Path | None]]:
    """Scan the data directory for dataset YAML files.

    Returns:
        Mapping of ``dataset_id`` -> ``{"codebook": Path|None,
        "meta": Path|None}``.
    """
    result: dict[str, dict[str, Path | None]] = {}
    for yml in DATA_DIR.glob("*_register_*.yaml"):
        dataset_id = _dataset_id_from_filename(yml)
        bucket = result.setdefault(
            dataset_id,
            {"codebook": None, "meta": None},
        )
        if yml.name.endswith("_register_codebook.yaml"):
            bucket["codebook"] = yml
        elif yml.name.endswith("_register_meta.yaml"):
            bucket["meta"] = yml
    return result


def _dataset_md_path_for(
    dataset_id: str,
    meta_path: Path | None,
) -> Path:
    """Derive the path to a dataset's Markdown description file.

    The Markdown file sits in ``./data`` alongside the YAML files and
    is named identically to ``*_register_meta.yaml`` but with a ``.md``
    extension.

    Args:
        dataset_id: Dataset identifier (e.g. ``ps_cancer``).
        meta_path: Path to the meta YAML, or ``None`` if absent.

    Returns:
        Expected path to the Markdown file (may not exist on disk).
    """
    if meta_path:
        return meta_path.with_suffix(".md")
    return DATA_DIR / f"{dataset_id}_register_meta.md"


def _load_dataset_markdown(
    dataset_id: str,
    meta_path: Path | None,
) -> tuple[str, Path | None]:
    """Load a dataset's Markdown description, if it exists.

    Args:
        dataset_id: Dataset identifier.
        meta_path: Path to the meta YAML, or ``None``.

    Returns:
        A 2-tuple of ``(markdown_text, md_path)`` where *md_path* is
        ``None`` when no file was found.
    """
    md_path = _dataset_md_path_for(dataset_id, meta_path)
    if md_path.exists() and md_path.is_file():
        return md_path.read_text(encoding="utf-8"), md_path
    return "", None


# ──────────────────────────────────────────────────────────────────────
# Variable normalisation
# ──────────────────────────────────────────────────────────────────────

def _normalize_var(raw_var: _YamlDict) -> _YamlDict:
    """Normalise a single variable dictionary.

    Codebook YAML files use varying field names for the same concept
    (e.g. ``labels`` vs ``label``, ``coltypes`` vs ``dtype``).  This
    function copies the input and adds canonical keys so downstream
    code can rely on a stable schema.

    Args:
        raw_var: Variable metadata dict read from YAML.

    Returns:
        A copy of *raw_var* with canonical keys added.
    """
    normalized = dict(raw_var)

    # Standardize name.
    if "name" not in normalized:
        normalized["name"] = (
            normalized.get("colname_silver")
            or normalized.get("colname")
            or None
        )

    # Normalise label — try several source fields.
    normalized["label"] = (
        normalized.get("labels")
        or normalized.get("label")
        or normalized.get("name")
        or ""
    )

    # Normalise type / data-type field.
    normalized["type"] = (
        normalized.get("coltypes")
        or normalized.get("dtype")
        or normalized.get("class")
        or ""
    )

    # Normalise categories and tags to lists.
    categories = normalized.get("categories")
    if isinstance(categories, str):
        categories = [categories]
    normalized["categories"] = categories or []

    tags = normalized.get("tags")
    if isinstance(tags, str):
        tags = [tags]
    normalized["tags"] = tags or []

    # Other fields.
    normalized["notes"] = normalized.get("notes") or ""
    # ``source`` = original column name from the upstream register.
    normalized["source"] = (
        normalized.get("colname_silver")
        or normalized.get("source")
        or ""
    )
    normalized["is_group"] = bool(normalized.get("is_group", False))

    return normalized


def _normalize_var_map(raw_map: dict[str, _YamlDict]) -> dict[str, _YamlDict]:
    """Normalise every variable in a name -> metadata mapping.

    Args:
        raw_map: Raw mapping from variable name to metadata dict.

    Returns:
        A new mapping with the same keys and normalised metadata.
    """
    normalized: dict[str, _YamlDict] = {}
    for name, var_info in raw_map.items():
        if not isinstance(var_info, dict):
            continue
        var_info = dict(var_info)
        var_info.setdefault("name", name)
        normalized[name] = _normalize_var(var_info)
    return normalized


def _extract_var_map_from_codebook(
    codebook_data: Any,
) -> dict[str, _YamlDict]:
    """Parse a codebook YAML into a normalised variable mapping.

    Supports several YAML shapes produced by different upstream tools:

    * ``{variables: [{name: ..., ...}, ...]}``  (list of dicts)
    * ``{variables: {varname: {...}, ...}}``     (nested mapping)
    * ``{var_map: {varname: {...}, ...}}``       (alternative key)
    * ``{varname: {...}, ...}``                  (flat mapping)

    Args:
        codebook_data: Parsed YAML content from a codebook file.

    Returns:
        Normalised mapping of variable name -> metadata dict.
    """
    if not codebook_data:
        return {}

    if isinstance(codebook_data, dict):
        if "var_map" in codebook_data and isinstance(
            codebook_data["var_map"], dict,
        ):
            return _normalize_var_map(codebook_data["var_map"])

        if "variables" in codebook_data:
            vars_obj = codebook_data["variables"]
            if isinstance(vars_obj, list):
                return {
                    v.get("name"): _normalize_var(v)
                    for v in vars_obj
                    if v and v.get("name")
                }
            if isinstance(vars_obj, dict):
                return _normalize_var_map(vars_obj)

        if all(
            isinstance(v, dict) for v in codebook_data.values()
        ):
            return _normalize_var_map(codebook_data)

    return {}


# ──────────────────────────────────────────────────────────────────────
# Filtering
# ──────────────────────────────────────────────────────────────────────

def _apply_ignore(
    var_map: dict[str, _YamlDict],
    ignore_names: list[str],
    ignore_name_patterns: list[re.Pattern[str]],
    ignore_tags: list[str] | None = None,
    ignore_categories: list[str] | None = None,
) -> dict[str, _YamlDict]:
    """Remove variables that match any ignore rule.

    A variable is removed if **any** of the following hold:

    * Its name appears in *ignore_names*.
    * Its name matches any pattern in *ignore_name_patterns*.
    * It carries a tag listed in *ignore_tags*.
    * It belongs to a category listed in *ignore_categories*.

    Args:
        var_map: Variable name -> metadata mapping.
        ignore_names: Exact variable names to remove.
        ignore_name_patterns: Compiled regex patterns.
        ignore_tags: Tag strings to match against variable tags.
        ignore_categories: Category strings to match.

    Returns:
        A new mapping containing only the variables that survived
        filtering.
    """
    if ignore_tags is None:
        ignore_tags = []
    if ignore_categories is None:
        ignore_categories = []

    filtered: dict[str, _YamlDict] = {}
    for name, var_info in var_map.items():
        if name in ignore_names:
            continue
        if any(p.search(name) for p in ignore_name_patterns):
            continue
        var_tags = var_info.get("tags") or []
        if any(tag in ignore_tags for tag in var_tags):
            continue
        var_cats = var_info.get("categories") or []
        if any(cat in ignore_categories for cat in var_cats):
            continue
        filtered[name] = var_info
    return filtered


# ──────────────────────────────────────────────────────────────────────
# Grouping
# ──────────────────────────────────────────────────────────────────────

def _build_groups_for_dataset(
    groups_config: dict[str, _YamlDict],
    var_map: dict[str, _YamlDict],
) -> list[_YamlDict]:
    """Create synthetic group variables from config patterns.

    Each group definition supplies a regex ``pattern``.  Variables whose
    names match that pattern are collapsed into a single "group" row in
    the final listing.  The group's categories are derived from its
    members' categories according to one of three strategies:

    * **union** (default): all categories found in any member.
    * **intersection**: only categories shared by every member.
    * **override**: a fixed list supplied in ``categories_override``.

    Args:
        groups_config: Mapping of group-ID -> group definition dict.
        var_map: Current variable mapping (post-ignore filtering).

    Returns:
        A list of synthetic group dicts sorted by
        ``(priority, label)``.
    """
    groups: list[_YamlDict] = []

    for group_id, group_def in (groups_config or {}).items():
        pattern = group_def.get("pattern")
        if not pattern:
            continue
        try:
            regex = re.compile(pattern)
        except re.error:
            logger.warning(
                "Invalid regex '%s' for group '%s'; skipping.",
                pattern,
                group_id,
            )
            continue

        members = sorted(
            name for name in var_map if regex.search(name)
        )
        if not members:
            continue

        # Derive categories from members according to the chosen
        # strategy (union / intersection / override).
        strategy = (
            group_def.get("category_strategy") or "union"
        ).lower()
        categories_override = (
            group_def.get("categories_override") or []
        )

        if strategy == "override":
            categories = list(categories_override)
        else:
            member_cats = [
                set(var_map[m].get("categories") or [])
                for m in members
            ]
            if not member_cats:
                categories = []
            elif strategy == "intersection":
                categories = sorted(
                    set.intersection(*member_cats),
                )
            else:  # "union" (default)
                categories = sorted(
                    set().union(*member_cats),
                )

        group_var: _YamlDict = {
            "name": group_id,
            "label": group_def.get("label") or group_id,
            "is_group": True,
            "members": members,
            "notes": group_def.get("notes") or "",
            "source": (
                group_def.get("source_variable_name_grouped")
                or ""
            ),
            "type": "",
            "categories": categories,
            "tags": [],
            # Client-side behaviour hints.
            "csv_expand": group_def.get("csv_expand") or "",
            "category_strategy": (
                group_def.get("category_strategy") or ""
            ),
            "categories_override": (
                group_def.get("categories_override") or []
            ),
            "_priority": (
                int(group_def.get("priority", 1000))
                if "priority" in group_def
                else 1000
            ),
        }
        groups.append(group_var)

    groups.sort(
        key=lambda g: (
            g.get("_priority", 1000),
            (g.get("label", g["name"]) or "").lower(),
        ),
    )
    return groups


# ──────────────────────────────────────────────────────────────────────
# Notes preview
# ──────────────────────────────────────────────────────────────────────

def _notes_preview_and_flag(
    text: str | None,
) -> tuple[str, bool]:
    """Build a short preview of a notes field.

    The preview collapses all internal whitespace.  If the original
    text is longer than the preview (i.e. it was truncated), the second
    element of the returned tuple is ``True``, signalling that the
    client should render an expand/collapse toggle.

    Heuristics:

    * If the text contains a double newline, the first paragraph is
      used as the preview.
    * Otherwise, the text is hard-capped at 120 characters (breaking
      at the nearest preceding word boundary).

    Args:
        text: Raw notes string (may be ``None`` or empty).

    Returns:
        A 2-tuple ``(preview, is_shortened)``.
    """
    if not text:
        return "", False

    raw = str(text)
    collapsed_full = re.sub(r"\s+", " ", raw).strip()

    # Choose a snippet: prefer the first paragraph when double-newline
    # separators exist; otherwise hard-cap at 120 characters.
    if "\n\n" in raw:
        snippet = raw.split("\n\n", 1)[0]
    else:
        max_len = 120
        if len(collapsed_full) <= max_len:
            snippet = collapsed_full
        else:
            cut = collapsed_full.rfind(" ", 0, max_len)
            if cut == -1:
                cut = max_len
            snippet = collapsed_full[:cut]

    preview = re.sub(r"\s+", " ", snippet).strip()
    shortened = preview != collapsed_full
    if shortened and not preview.endswith("\u2026"):
        preview = preview.rstrip(" .") + "\u2026"
    return preview, shortened


# ──────────────────────────────────────────────────────────────────────
# Dataset assembly
# ──────────────────────────────────────────────────────────────────────

def _assemble_dataset(
    dataset_id: str,
    codebook_path: Path | None,
    meta_path: Path | None,
    groups_config: dict[str, _YamlDict],
    ignore_names: list[str],
    ignore_name_patterns: list[re.Pattern[str]],
    ignore_tags: list[str],
    ignore_categories: list[str],
) -> tuple[_YamlDict, list[tuple[str, str]]]:
    """Load, normalise, filter and group a single dataset.

    This is the main per-dataset pipeline entry point.  It:

    1. Loads the meta YAML and codebook YAML.
    2. Normalises variable metadata.
    3. Applies ignore rules.
    4. Synthesises group variables.
    5. Builds the final ordered variable list.
    6. Annotates each variable with a notes preview and flag.
    7. Loads the Markdown dataset description (or YAML fallback).

    Args:
        dataset_id: Short identifier (e.g. ``ps_cancer``).
        codebook_path: Path to the codebook YAML, or ``None``.
        meta_path: Path to the meta YAML, or ``None``.
        groups_config: Group definitions from variables_config.
        ignore_names: Exact variable names to remove.
        ignore_name_patterns: Compiled regex patterns for names.
        ignore_tags: Tags that trigger removal.
        ignore_categories: Categories that trigger removal.

    Returns:
        A 2-tuple ``(dataset_dict, provenance_entries)`` where
        *provenance_entries* is a list of ``(filename, sha256_hex)``
        pairs for every input file consumed.
    """
    provenance: list[tuple[str, str]] = []

    # ── 1. Load meta YAML (optional). ────────────────────────────
    meta: _YamlDict = {}
    if meta_path and meta_path.exists():
        meta = _read_yaml(meta_path) or {}
        provenance.append(
            (meta_path.name, _sha256_file(meta_path)),
        )

    # ── 2. Load and normalise codebook. ──────────────────────────
    var_map: dict[str, _YamlDict] = {}
    if codebook_path and codebook_path.exists():
        codebook_data = _read_yaml(codebook_path)
        var_map = _extract_var_map_from_codebook(codebook_data)
        for var_info in var_map.values():
            var_info["_source_yaml"] = codebook_path.name
        provenance.append(
            (codebook_path.name, _sha256_file(codebook_path)),
        )

    # ── 3. Apply ignore rules. ───────────────────────────────────
    var_map = _apply_ignore(
        var_map,
        ignore_names,
        ignore_name_patterns,
        ignore_tags,
        ignore_categories,
    )

    # Keep a full copy *after* ignore rules.  Individual members may
    # be hidden from the visible listing when a synthetic group
    # subsumes them, but the full map is still needed for client-side
    # CSV export where groups expand back into their members.
    var_map_all = dict(var_map)

    # ── 4. Build synthetic groups. ───────────────────────────────
    groups = _build_groups_for_dataset(groups_config, var_map)

    # Remove individual members that are now represented by a group,
    # keeping the listing concise and avoiding duplicated rows.
    grouped_members: set[str] = set()
    for group in groups:
        for member in group.get("members", []):
            grouped_members.add(member)
    for member in grouped_members:
        var_map.pop(member, None)

    # ── 5. Ordered variable list. ────────────────────────────────
    # Place each synthetic group at the position of its first member
    # in the original codebook order so the listing stays intuitive.
    variables: list[_YamlDict] = []

    # Map member -> owning group (first group by priority wins).
    member_to_group: dict[str, _YamlDict] = {}
    for group in groups:
        for member in group.get("members", []):
            if member not in member_to_group:
                member_to_group[member] = group

    inserted_groups: set[str] = set()
    for name in list(var_map_all.keys()):
        if name in member_to_group:
            group = member_to_group[name]
            if group["name"] not in inserted_groups:
                variables.append(group)
                inserted_groups.add(group["name"])
            continue
        if name in var_map:
            variables.append(var_map[name])

    # Append groups whose members weren't present in var_map_all.
    for group in groups:
        if group["name"] not in inserted_groups:
            variables.append(group)

    # Defensive: append orphan variables not yet in the list.
    for var_info in var_map.values():
        if var_info not in variables:
            variables.append(var_info)

    # ── 6. Annotate notes previews. ──────────────────────────────
    for var_info in var_map_all.values():
        preview, shortened = _notes_preview_and_flag(
            var_info.get("notes"),
        )
        var_info["notes_preview"] = preview
        var_info["notes_is_long"] = bool(shortened)

    for group in groups:
        preview, shortened = _notes_preview_and_flag(
            group.get("notes"),
        )
        group["notes_preview"] = preview
        group["notes_is_long"] = bool(shortened)

    for var_info in var_map.values():
        preview, shortened = _notes_preview_and_flag(
            var_info.get("notes"),
        )
        var_info["notes_preview"] = preview
        var_info["notes_is_long"] = bool(shortened)

    # Propagate flags into the merged ``variables`` list.
    for entry in variables:
        if not isinstance(entry, dict):
            continue
        entry_name: Any = entry.get("name")
        if entry_name and entry_name in var_map_all:
            source = var_map_all[entry_name]
            entry["notes_is_long"] = source.get(
                "notes_is_long", False,
            )
            entry["notes_preview"] = source.get(
                "notes_preview",
                source.get("notes", "") or "",
            )
        else:
            entry.setdefault("notes_is_long", False)
            entry.setdefault(
                "notes_preview",
                entry.get("notes", "") or "",
            )

    # ── 7. Dataset description (Markdown preferred). ─────────────
    info_md, md_path = _load_dataset_markdown(
        dataset_id, meta_path,
    )
    if md_path:
        provenance.append(
            (md_path.name, _sha256_file(md_path)),
        )

    if info_md.strip():
        info_block: _YamlDict = {
            "info_html": _md_to_html(
                info_md, preserve_newlines=True,
            ),
        }
    else:
        yaml_info = (
            meta.get("info")
            if isinstance(meta.get("info"), list)
            else []
        )
        info_block = {"info": yaml_info}

    dataset = {
        "id": dataset_id,
        "title": (
            meta.get("title")
            or _readable_title_from_id(dataset_id)
        ),
        "subtitle": meta.get("subtitle") or "",
        **info_block,
        "var_map": var_map,
        "var_map_all": var_map_all,
        "variables": variables,
    }
    return dataset, provenance


def _readable_title_from_id(dataset_id: str) -> str:
    """Generate a human-readable title from a dataset identifier.

    Args:
        dataset_id: e.g. ``ps_cancer``.

    Returns:
        Title-cased string, e.g. ``Ps Cancer``.
    """
    return (
        dataset_id.replace("_", " ").replace("-", " ").title()
    )


# ──────────────────────────────────────────────────────────────────────
# Content loaders (intro, footer, logo)
# ──────────────────────────────────────────────────────────────────────

def _load_intro_html() -> str:
    """Load and render ``content/intro.md`` to HTML.

    Returns:
        Rendered HTML string, or ``""`` if the file is absent.
    """
    if INTRO_MD.exists():
        return _md_to_html(INTRO_MD.read_text(encoding="utf-8"))
    return ""


def _load_footer_html() -> str:
    """Load and render ``content/footer.md`` to HTML.

    If the footer markdown contains the ``[LOGO]`` placeholder and a
    logo image exists in ``content/``, the placeholder is replaced
    with an embedded ``<img>`` tag.

    Returns:
        Rendered HTML string, or ``""`` if the file is absent.
    """
    if not FOOTER_MD.exists():
        return ""
    html = _md_to_html(FOOTER_MD.read_text(encoding="utf-8"))
    logo_uri = _load_logo_data_uri()
    if logo_uri and "[LOGO]" in html:
        logo_tag = (
            f'<img src="{logo_uri}" alt="Logo" '
            f'style="height:64px;width:auto;'
            f'display:block;margin-bottom:1rem;" />'
        )
        html = html.replace("[LOGO]", logo_tag)
    return html


def _load_logo_data_uri() -> str:
    """Encode the logo image as a data-URI for embedding in HTML.

    Supports PNG, JPEG, and SVG formats.

    Returns:
        A ``data:`` URI string, or ``""`` if no logo file exists.
    """
    if not LOGO_PATH.exists():
        return ""
    data = LOGO_PATH.read_bytes()
    suffix = LOGO_PATH.suffix.lower()

    if suffix == ".png":
        mime = "image/png"
    elif suffix in (".jpg", ".jpeg"):
        mime = "image/jpeg"
    elif suffix == ".svg":
        try:
            text = data.decode("utf-8")
            return f"data:image/svg+xml;utf8,{text}"
        except Exception:
            mime = "image/svg+xml"
    else:
        mime = "application/octet-stream"

    b64 = base64.b64encode(data).decode("ascii")
    return f"data:{mime};base64,{b64}"


# ──────────────────────────────────────────────────────────────────────
# Provenance
# ──────────────────────────────────────────────────────────────────────

def _format_provenance(
    provenance_per_dataset: dict[str, list[tuple[str, str]]],
) -> str:
    """Format a provenance manifest from per-dataset hash lists.

    The manifest lists each input file and its SHA-256 digest, followed
    by a combined digest computed over all individual hex digests in
    stable order.

    Args:
        provenance_per_dataset: Mapping of dataset-ID -> list of
            ``(filename, sha256_hex)`` pairs.

    Returns:
        Multi-line string suitable for embedding in the output HTML.
    """
    lines: list[str] = []
    all_hexes: list[str] = []

    # Supplemental config / content files.
    supplemental_files: list[Path] = [
        p
        for p in (VARIABLES_CONFIG, INTRO_MD, FOOTER_MD, LOGO_PATH)
        if p.exists()
    ]

    # Per-dataset YAMLs and MDs.
    for dataset_id in sorted(provenance_per_dataset):
        for filename, digest in provenance_per_dataset[dataset_id]:
            lines.append(f"{filename} {digest}")
            all_hexes.append(digest)

    # Supplemental files.
    for supplemental_path in supplemental_files:
        digest = _sha256_file(supplemental_path)
        lines.append(f"{supplemental_path.name} {digest}")
        all_hexes.append(digest)

    if all_hexes:
        combined = _sha256_bytes(
            "".join(all_hexes).encode("utf-8"),
        )
        lines.append(f"Combined SHA-256: {combined}")

    return "\n".join(lines)


# ──────────────────────────────────────────────────────────────────────
# Dataset-group ordering
# ──────────────────────────────────────────────────────────────────────

def _load_datasets_config() -> list[_YamlDict]:
    """Load display-group ordering from ``datasets_config.yaml``.

    Returns:
        A list of dicts, each with ``heading`` (str) and ``datasets``
        (list of dataset-ID strings).  Returns ``[]`` if the config
        file is missing or malformed.
    """
    if not DATASETS_CONFIG.exists():
        return []
    config = _read_yaml(DATASETS_CONFIG) or {}
    raw_groups = config.get("groups") or []

    result: list[_YamlDict] = []
    if isinstance(raw_groups, list):
        for group_entry in raw_groups:
            if not isinstance(group_entry, dict):
                continue
            heading = (
                group_entry.get("heading")
                or group_entry.get("label")
                or ""
            )
            datasets = (
                group_entry.get("datasets")
                or group_entry.get("ids")
                or []
            )
            if isinstance(datasets, str):
                datasets = [datasets]
            if not isinstance(datasets, list):
                datasets = []
            result.append({
                "heading": heading,
                "datasets": list(datasets),
            })
    return result


# ──────────────────────────────────────────────────────────────────────
# Template rendering
# ──────────────────────────────────────────────────────────────────────

def _get_git_sha() -> str:
    """Return the current ``HEAD`` commit SHA, or ``""`` on failure.

    Fails gracefully when run outside a Git repository or when the
    ``git`` binary is not available.
    """
    try:
        result = subprocess.run(
            ["git", "rev-parse", "HEAD"],
            cwd=str(ROOT_DIR),
            capture_output=True,
            text=True,
            timeout=5,
        )
        if result.returncode == 0:
            return result.stdout.strip()
    except Exception:
        logger.debug("Could not read git HEAD.", exc_info=True)
    return ""


def _get_build_date(
    input_files: list[Path] | None = None,
) -> str:
    """Determine the build timestamp in a reproducible way.

    Resolution order:

    1. **``SOURCE_DATE_EPOCH``** environment variable (integer Unix
       timestamp).  This is the `reproducible-builds.org`_ standard
       and is the recommended mechanism for CI / release builds.
    2. **Newest input-file mtime** — if *input_files* is supplied and
       non-empty, use the most recent modification time.
    3. **Current wall-clock time** (fallback for local development).

    .. _reproducible-builds.org:
       https://reproducible-builds.org/docs/source-date-epoch/

    Args:
        input_files: Optional list of paths whose mtimes should be
            considered for deterministic timestamps.

    Returns:
        ISO-8601 datetime string (whole seconds, UTC when derived
        from ``SOURCE_DATE_EPOCH``).
    """
    epoch = os.environ.get("SOURCE_DATE_EPOCH")
    if epoch is not None:
        try:
            return (
                datetime.fromtimestamp(
                    int(epoch), tz=timezone.utc,
                )
                .replace(microsecond=0)
                .isoformat()
            )
        except (ValueError, OSError, OverflowError):
            logger.warning(
                "Invalid SOURCE_DATE_EPOCH='%s'; ignoring.",
                epoch,
            )

    if input_files:
        try:
            newest = max(
                p.stat().st_mtime
                for p in input_files
                if p.exists()
            )
            return (
                datetime.fromtimestamp(newest)
                .replace(microsecond=0)
                .isoformat()
            )
        except (ValueError, OSError):
            pass

    return datetime.now().replace(microsecond=0).isoformat()


def _render_html(
    datasets: list[_YamlDict],
    intro_html: str,
    provenance_text: str,
    build_date: str,
) -> str:
    """Render the Jinja2 template to a complete HTML string.

    Args:
        datasets: Assembled dataset list (grouped or flat).
        intro_html: Pre-rendered HTML for the introduction section.
        provenance_text: Multi-line provenance manifest.
        build_date: ISO-8601 build timestamp.

    Returns:
        Complete HTML string ready for writing to disk.
    """
    env = Environment(
        loader=FileSystemLoader(str(TEMPLATES_DIR)),
        autoescape=select_autoescape(["html", "xml"]),
    )
    template = env.get_template(TEMPLATE_NAME)
    data_json = json.dumps(
        datasets, ensure_ascii=False, separators=(",", ":"),
    )

    return template.render(
        data_json=data_json,
        intro_html=intro_html,
        git_sha=_get_git_sha(),
        build_date=build_date,
        footer_html=_load_footer_html(),
        logo_data_uri=_load_logo_data_uri(),
    )


# ──────────────────────────────────────────────────────────────────────
# Excel export
# ──────────────────────────────────────────────────────────────────────

def _sanitize_sheet_name(name: str) -> str:
    r"""Sanitize a string for use as an Excel sheet name.

    Excel forbids ``[ ] : * ? / \`` and limits names to 31 chars.

    Args:
        name: Proposed sheet name.

    Returns:
        Sanitised name (max 31 chars, problematic chars removed).
    """
    sanitized = re.sub(r"[\[\]:*?/\\]", "", name)
    sanitized = sanitized[:31].strip()
    return sanitized or "Sheet"


def _auto_adjust_column_widths(worksheet: Worksheet) -> None:
    """Resize columns to fit their content (capped at 50 chars).

    Args:
        worksheet: An ``openpyxl`` worksheet object.
    """
    for column in worksheet.columns:
        max_length = 0
        col_letter = get_column_letter(
            column[0].column or 1,
        )
        for cell in column:
            try:
                if cell.value:
                    max_length = max(
                        max_length, len(str(cell.value)),
                    )
            except (TypeError, ValueError):
                pass
        worksheet.column_dimensions[col_letter].width = min(
            max_length + 2, 50,
        )


def _patch_xlsx_modified(xlsx_path: Path, stamp: datetime) -> None:
    """Rewrite the ``modified`` timestamp inside a saved XLSX file.

    openpyxl unconditionally sets ``modified`` to ``utcnow()`` in its
    ``save_workbook`` helper.  This function opens the XLSX (which is a
    ZIP archive), updates ``docProps/core.xml`` with *stamp*, and
    writes the archive back in place.

    Args:
        xlsx_path: Path to the ``.xlsx`` file.
        stamp: The datetime to use for the ``modified`` field.
    """
    import re as _re
    import zipfile

    iso = stamp.strftime("%Y-%m-%dT%H:%M:%S")
    if stamp.tzinfo is not None:
        iso += "Z"

    # Read original archive members.
    members: dict[str, bytes] = {}
    with zipfile.ZipFile(xlsx_path, "r") as zin:
        for info in zin.infolist():
            members[info.filename] = zin.read(info.filename)

    # Patch core.xml.
    core_key = "docProps/core.xml"
    if core_key in members:
        xml_text = members[core_key].decode("utf-8")
        xml_text = _re.sub(
            r"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
            rf"\g<1>{iso}\g<2>",
            xml_text,
        )
        members[core_key] = xml_text.encode("utf-8")

    # Rewrite the archive.
    with zipfile.ZipFile(xlsx_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in members.items():
            zout.writestr(name, data)


def _export_to_excel(
    datasets_raw: list[_YamlDict],
    output_path: Path,
    build_date: str = "",
) -> None:
    """Export datasets to a formatted Excel workbook.

    Each dataset occupies its own sheet.  Columns:

    ===================== =============================================
    Selection             Empty; users mark chosen variables with "x".
    Label                 Human-readable variable name.
    Notes                 Variable description.
    Variable name         YAML key (unique identifier).
    Source variable name  Original database variable name.
    Categories            Comma-separated list.
    Grouped               ``Yes`` / ``No``.
    Members               For groups: member variable names.
    ===================== =============================================

    Args:
        datasets_raw: Assembled datasets list (grouped or flat).
        output_path: Destination ``.xlsx`` path.
        build_date: Deterministic date string (``YYYY-MM-DD``) used to
            pin the workbook's ``created`` and ``modified`` metadata.
    """
    if not datasets_raw:
        logger.warning("No datasets to export to Excel.")
        return

    workbook = Workbook()
    default_sheet = workbook.active
    if default_sheet is not None:
        workbook.remove(default_sheet)

    # Pin document timestamps for reproducible builds.  openpyxl's
    # save_workbook() unconditionally overwrites ``modified`` with
    # utcnow(), so we keep a reference and restore it after save.
    pinned_stamp: datetime | None = None
    if build_date:
        pinned_stamp = datetime.fromisoformat(build_date)
        workbook.properties = DocumentProperties(
            created=pinned_stamp,
            modified=pinned_stamp,
        )

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="366092",
        end_color="366092",
        fill_type="solid",
    )
    header_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )
    cell_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Flatten grouped structure into a list of dataset dicts.
    flat_datasets: list[_YamlDict] = []
    for item in datasets_raw:
        if not isinstance(item, dict):
            continue
        if "heading" in item and "datasets" in item:
            for dataset_entry in item.get("datasets", []):
                if isinstance(dataset_entry, dict):
                    flat_datasets.append(dataset_entry)
        elif "id" in item and "variables" in item:
            flat_datasets.append(item)

    headers = [
        "Selection",
        "Label",
        "Notes",
        "Variable name",
        "Source variable name",
        "Categories",
        "Grouped",
        "Members",
    ]

    for dataset in flat_datasets:
        title: str = str(
            dataset.get("title", dataset.get("id", "unknown"))
        )
        variables = dataset.get("variables", [])
        worksheet = workbook.create_sheet(
            title=_sanitize_sheet_name(title),
        )

        # Write header row.
        for col_num, header_text in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header_text
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = cell_border

        # Write variable rows.
        for row_num, var_info in enumerate(variables, 2):
            if not isinstance(var_info, dict):
                continue

            var_categories = var_info.get("categories", [])
            members = var_info.get("members", [])
            row_data = [
                "",  # Selection (left blank for user input)
                var_info.get("label", ""),
                var_info.get("notes", ""),
                var_info.get("name", ""),
                var_info.get("source", ""),
                (
                    ", ".join(var_categories)
                    if isinstance(var_categories, list)
                    else str(var_categories)
                ),
                (
                    "Yes"
                    if var_info.get("is_group", False)
                    else "No"
                ),
                (
                    ", ".join(members)
                    if isinstance(members, list)
                    else str(members)
                ),
            ]

            for col_num, cell_value in enumerate(row_data, 1):
                cell = worksheet.cell(
                    row=row_num, column=col_num,
                )
                cell.value = cell_value
                cell.alignment = Alignment(
                    vertical="top", wrap_text=True,
                )
                cell.border = cell_border

        _auto_adjust_column_widths(worksheet)

    workbook.save(output_path)
    # Restore the pinned ``modified`` timestamp that openpyxl
    # overwrites during save, then re-serialise core.xml in-place.
    if pinned_stamp is not None:
        _patch_xlsx_modified(output_path, pinned_stamp)
    logger.info("Wrote %s", output_path)


# ──────────────────────────────────────────────────────────────────────
# Static file copy
# ──────────────────────────────────────────────────────────────────────

def _copy_static_files(dist_dir: Path) -> None:
    """Copy static assets (e.g. Word documents) to the output dir.

    Args:
        dist_dir: Destination directory (typically ``dist/``).
    """
    if not STATIC_DIR.exists():
        return
    for src in STATIC_DIR.iterdir():
        if src.is_file():
            shutil.copy2(src, dist_dir / src.name)
            logger.info(
                "Copied %s to %s/", src.name, dist_dir.name,
            )


# ──────────────────────────────────────────────────────────────────────
# Build entry point
# ──────────────────────────────────────────────────────────────────────

def build(output_dir: Path | None = None) -> None:
    """Run the full build pipeline.

    1. Load configuration and discover datasets.
    2. Assemble each dataset (normalise, filter, group).
    3. Render HTML and export Excel.
    4. Copy static files to the output directory.

    Args:
        output_dir: Override the default output directory
            (``./dist``).  Created automatically if absent.
    """
    dist_dir = output_dir or DIST_DIR
    dist_dir.mkdir(parents=True, exist_ok=True)

    (
        groups_config,
        ignore_names,
        ignore_name_patterns,
        ignore_tags,
        ignore_categories,
    ) = _load_variables_config()

    discovered = _discover_datasets()
    datasets: list[_YamlDict] = []
    provenance: dict[str, list[tuple[str, str]]] = {}

    # Collect every input file path for deterministic timestamping.
    all_input_files: list[Path] = []

    ds_groups_cfg = _load_datasets_config()
    remaining = set(discovered.keys())

    def _assemble_and_record(
        dataset_id: str,
    ) -> _YamlDict:
        """Assemble one dataset and record its provenance."""
        paths = discovered[dataset_id]
        dataset, dataset_prov = _assemble_dataset(
            dataset_id=dataset_id,
            codebook_path=paths.get("codebook"),
            meta_path=paths.get("meta"),
            groups_config=groups_config,
            ignore_names=ignore_names,
            ignore_name_patterns=ignore_name_patterns,
            ignore_tags=ignore_tags,
            ignore_categories=ignore_categories,
        )
        provenance[dataset_id] = dataset_prov
        remaining.discard(dataset_id)
        # Track input files for timestamp derivation.
        for key in ("codebook", "meta"):
            input_path = paths.get(key)
            if input_path:
                all_input_files.append(input_path)
        md_path = _dataset_md_path_for(
            dataset_id, paths.get("meta"),
        )
        if md_path.exists():
            all_input_files.append(md_path)
        return dataset

    if ds_groups_cfg:
        for group in ds_groups_cfg:
            heading = group.get("heading") or ""
            group_list: list[_YamlDict] = []
            for dataset_id in group.get("datasets") or []:
                if dataset_id not in discovered:
                    logger.warning(
                        "Dataset '%s' listed in %s not "
                        "found in data directory; skipping.",
                        dataset_id,
                        DATASETS_CONFIG.name,
                    )
                    continue
                group_list.append(
                    _assemble_and_record(dataset_id),
                )
            if group_list:
                datasets.append({
                    "heading": heading,
                    "datasets": group_list,
                })

    # Append remaining (ungrouped) datasets alphabetically.
    if remaining:
        others = [
            _assemble_and_record(dataset_id)
            for dataset_id in sorted(remaining)
        ]
        if ds_groups_cfg:
            datasets.append({
                "heading": "Other datasets",
                "datasets": others,
            })
        else:
            datasets = others

    # Add supplemental config / content files to the input list.
    for path in (VARIABLES_CONFIG, INTRO_MD, FOOTER_MD, LOGO_PATH):
        if path.exists():
            all_input_files.append(path)

    intro_html = _load_intro_html()
    excel_filename = "variables_browser.xlsx"
    if intro_html and "[EXCEL-SELECTIONS]" in intro_html:
        intro_html = intro_html.replace(
            "[EXCEL-SELECTIONS]",
            f'<a href="{excel_filename}">Download Excel file</a>',
        )

    provenance_text = _format_provenance(provenance)
    build_date = _get_build_date(all_input_files)

    html = _render_html(
        datasets, intro_html, provenance_text, build_date,
    )
    output_html = dist_dir / OUTPUT_HTML_NAME
    output_html.write_text(html, encoding="utf-8")
    logger.info("Wrote %s", output_html)

    _export_to_excel(datasets, dist_dir / excel_filename, build_date)
    _copy_static_files(dist_dir)


# ──────────────────────────────────────────────────────────────────────
# CLI
# ──────────────────────────────────────────────────────────────────────

def _parse_args() -> argparse.Namespace:
    """Parse command-line arguments.

    Returns:
        Parsed :class:`argparse.Namespace`.
    """
    parser = argparse.ArgumentParser(
        description=(
            "Build the PREDICT data-portal HTML and Excel files."
        ),
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=None,
        help=(
            "Output directory (default: ./dist). "
            "Created automatically if absent."
        ),
    )
    verbosity = parser.add_mutually_exclusive_group()
    verbosity.add_argument(
        "--verbose", "-v",
        action="store_true",
        help="Enable debug-level logging.",
    )
    verbosity.add_argument(
        "--quiet", "-q",
        action="store_true",
        help="Suppress info messages; show warnings only.",
    )
    return parser.parse_args()


def main() -> None:
    """CLI entry point: parse arguments, configure logging, build."""
    args = _parse_args()

    if args.verbose:
        level = logging.DEBUG
    elif args.quiet:
        level = logging.WARNING
    else:
        level = logging.INFO

    logging.basicConfig(
        format="%(levelname)s: %(message)s",
        level=level,
    )

    build(output_dir=args.output_dir)


if __name__ == "__main__":
    main()
